from datetime import date

from openpyxl import load_workbook

from ape.database.database import DatabaseManager
from ape.database.models import Draw
from ape.database.repositories import DrawRepository
from ape.reports.excel_exporter import ExcelReportExporter


def make_draw(row_date: date, values: list[int]) -> Draw:
    values = sorted(values)
    odd_count = sum(value % 2 for value in values)
    low_count = sum(value <= 22 for value in values)
    return Draw(
        draw_date=row_date,
        weekday_index=row_date.weekday(),
        weekday_name=(
            "Thứ Hai",
            "Thứ Ba",
            "Thứ Tư",
            "Thứ Năm",
            "Thứ Sáu",
            "Thứ Bảy",
            "Chủ Nhật",
        )[row_date.weekday()],
        n1=values[0],
        n2=values[1],
        n3=values[2],
        n4=values[3],
        n5=values[4],
        n6=values[5],
        total_sum=sum(values),
        odd_count=odd_count,
        even_count=6 - odd_count,
        low_count=low_count,
        high_count=6 - low_count,
        source_file="test.xlsx",
        source_row=2,
    )


def test_excel_report_exporter_creates_workbook(tmp_path):
    database = DatabaseManager(tmp_path / "ape.db")
    database.initialize()
    rows = [
        make_draw(date(2026, 1, 1), [1, 2, 3, 4, 5, 6]),
        make_draw(date(2026, 1, 3), [1, 2, 7, 8, 9, 10]),
        make_draw(date(2026, 1, 5), [1, 3, 7, 11, 12, 13]),
    ]
    with database.session() as session:
        repository = DrawRepository(session)
        for row in rows:
            repository.add(row)

    output_file = ExcelReportExporter(database).export(tmp_path / "report.xlsx")
    workbook = load_workbook(output_file)

    assert output_file.exists()
    assert "Tong_quan" in workbook.sheetnames
    assert "Du_lieu" in workbook.sheetnames
    assert "Thong_ke_01_45" in workbook.sheetnames
    assert "Bieu_do" in workbook.sheetnames
    assert workbook["Tong_quan"]["B3"].value == 3

    database.dispose()
