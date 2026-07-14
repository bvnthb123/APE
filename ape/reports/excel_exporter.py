"""Excel report exporter for APE historical analysis."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ape.analytics.service import AnalysisService
from ape.core.settings import SETTINGS
from ape.database.database import DATABASE, DatabaseManager
from ape.database.repositories import DrawRepository

HEADER_FILL = PatternFill("solid", fgColor="005BAC")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(color="005BAC", bold=True, size=14)


class ExcelReportExporter:
    """Export database history and descriptive statistics to Excel."""

    def __init__(self, database: DatabaseManager | None = None) -> None:
        self.database = database or DATABASE

    def export(
        self,
        output_path: Path | str | None = None,
        *,
        limit: int = 15,
    ) -> Path:
        self.database.initialize()
        with self.database.session() as session:
            draws = DrawRepository(session).list_chronological()

        report = AnalysisService(self.database).generate(limit=limit)
        target = self._resolve_output_path(output_path)

        workbook = Workbook()
        workbook.remove(workbook.active)

        self._summary_sheet(workbook, report)
        self._history_sheet(workbook, draws)
        self._number_metrics_sheet(workbook, report)
        self._groups_sheet(workbook, "Cap_so", report.common_pairs)
        self._groups_sheet(workbook, "Bo_ba", report.common_triples)
        self._audit_sheet(workbook, report)
        self._charts_sheet(workbook)

        workbook.save(target)
        return target

    def _resolve_output_path(self, output_path: Path | str | None) -> Path:
        if output_path:
            path = Path(output_path).expanduser().resolve()
        else:
            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            path = SETTINGS.reports_dir / f"ape_report_{stamp}.xlsx"
        path.parent.mkdir(parents=True, exist_ok=True)
        return path

    def _summary_sheet(self, workbook: Workbook, report) -> None:
        sheet = workbook.create_sheet("Tong_quan")
        rows = [
            ("APE Report", ""),
            ("Ngay xuat", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
            ("Tong so ky", report.dataset["total_rows"]),
            ("Ngay dau", report.dataset["first_date"]),
            ("Ngay cuoi", report.dataset["last_date"]),
            ("Chat luong du lieu", report.audit["quality_score"]),
            ("Dong loi", report.audit["invalid_row_count"]),
            ("Ngay trung", report.audit["duplicate_date_count"]),
            ("Sai thu", report.audit["weekday_mismatch_count"]),
        ]
        for row in rows:
            sheet.append(row)
        sheet["A1"].font = TITLE_FONT
        self._autosize(sheet)

    def _history_sheet(self, workbook: Workbook, draws: Iterable[Any]) -> None:
        sheet = workbook.create_sheet("Du_lieu")
        sheet.append([
            "Ngay",
            "Thu",
            "Bo so",
            "Tong",
            "Le",
            "Chan",
            "Thap",
            "Cao",
            "Nguon",
        ])
        for draw in draws:
            sheet.append([
                draw.draw_date.strftime("%d/%m/%Y"),
                draw.weekday_name,
                " - ".join(f"{number:02d}" for number in draw.numbers),
                draw.total_sum,
                draw.odd_count,
                draw.even_count,
                draw.low_count,
                draw.high_count,
                draw.source_file or "",
            ])
        self._style_table(sheet)

    def _number_metrics_sheet(self, workbook: Workbook, report) -> None:
        sheet = workbook.create_sheet("Thong_ke_01_45")
        sheet.append([
            "Gia tri",
            "So lan",
            "Ty le",
            "Khoang vang hien tai",
            "Gap trung binh",
            "Gap lon nhat",
            "30 ky gan",
            "Xu huong",
        ])
        for item in report.event_metrics:
            sheet.append([
                item["event_id"],
                item["count"],
                item["rate"],
                item["latest_distance"],
                item["mean_distance"],
                item["largest_distance"],
                item["recent_count"],
                item["change_rate"],
            ])
        self._style_table(sheet)

    def _groups_sheet(self, workbook: Workbook, name: str, rows: list[dict[str, Any]]) -> None:
        sheet = workbook.create_sheet(name)
        sheet.append(["Nhom", "So lan", "Ty le"])
        for item in rows:
            sheet.append([item["values"], item["count"], item["rate"]])
        self._style_table(sheet)

    def _audit_sheet(self, workbook: Workbook, report) -> None:
        sheet = workbook.create_sheet("Kiem_tra")
        audit = report.audit
        rows = [
            ("Diem chat luong", audit["quality_score"]),
            ("Dong loi", audit["invalid_row_count"]),
            ("Ngay trung", audit["duplicate_date_count"]),
            ("Sai thu", audit["weekday_mismatch_count"]),
            ("Thieu thong tin nguon", audit["missing_source_metadata_count"]),
        ]
        for row in rows:
            sheet.append(row)
        sheet.append([])
        sheet.append(["Khoang cach ngay lon hon 14 ngay"])
        sheet.append(["Tu ngay", "Den ngay", "So ngay"])
        for item in audit.get("long_date_gaps_over_14_days", []):
            sheet.append([item["from"], item["to"], item["days"]])
        self._style_table(sheet)

    def _charts_sheet(self, workbook: Workbook) -> None:
        sheet = workbook.create_sheet("Bieu_do")
        data_sheet = workbook["Thong_ke_01_45"]

        chart = BarChart()
        chart.title = "Tan suat 01-45"
        chart.y_axis.title = "So lan"
        chart.x_axis.title = "Gia tri"
        data = Reference(data_sheet, min_col=2, min_row=1, max_row=46)
        categories = Reference(data_sheet, min_col=1, min_row=2, max_row=46)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 9
        chart.width = 22
        sheet.add_chart(chart, "A1")

        line_chart = LineChart()
        line_chart.title = "Khoang vang hien tai"
        line_chart.y_axis.title = "So ky"
        line_chart.x_axis.title = "Gia tri"
        data = Reference(data_sheet, min_col=4, min_row=1, max_row=46)
        line_chart.add_data(data, titles_from_data=True)
        line_chart.set_categories(categories)
        line_chart.height = 9
        line_chart.width = 22
        sheet.add_chart(line_chart, "A20")

    def _style_table(self, sheet) -> None:
        if sheet.max_row >= 1:
            for cell in sheet[1]:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center")
        self._autosize(sheet)

    @staticmethod
    def _autosize(sheet) -> None:
        for column_cells in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            sheet.column_dimensions[column_letter].width = min(max_length + 2, 36)
